from django.contrib import messages
from django.db import transaction
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.response import TemplateResponse

from apps.tenant_apps.utils.htmx_utils import for_htmx

from ..forms import LedgerForm, LedgerStatementForm, LedgerTransactionForm
from ..models import JournalEntry, Ledger, LedgerStatement, LedgerTransaction


def set_ledger_ob(request, pk):
    # if this is a POST request we need to process the form data
    if request.method == "POST":
        # create a form instance and populate it with data from the request:
        form = LedgerStatementForm(request.POST)
        # check whether it's valid:
        if form.is_valid():
            # process the data in form.cleaned_data as required
            # ...
            # redirect to a new URL:
            form.save()
            return redirect("/")

    # if a GET (or any other method) we'll create a blank form
    else:
        form = LedgerStatementForm()

    return render(request, "dea/set_ledger_ob.html", {"form": form})


@transaction.atomic()
def audit_ledger(request):
    ledgers = Ledger.objects.all()
    for l in ledgers:
        l.audit()
    return redirect("/dea")


@for_htmx(use_block="content")
def ledger_list(request):
    ledgers = Ledger.objects.all()
    return TemplateResponse(request, "dea/ledger_list.html", {"ledgers": ledgers})


def ledger_detail(request, pk):
    ledger = get_object_or_404(Ledger, id=pk)
    ls_created = (
        ledger.ledgerstatements.latest().created
        if ledger.ledgerstatements.exists()
        else None
    )
    dtxns = ledger.dtxns(since=ls_created).select_related("journal_entry__content_type")
    ctxns = ledger.ctxns(since=ls_created).select_related("journal_entry__content_type")
    cr_aleg_txns = ledger.aleg_txns(xacttypecode="Cr", since=ls_created)
    dr_aleg_txns = ledger.aleg_txns(xacttypecode="Dr", since=ls_created)
    return render(
        request,
        "dea/ledger_detail.html",
        {
            "object": ledger,
            "dtxns": dtxns,
            "ctxns": ctxns,
            "cr_aleg_txns": cr_aleg_txns,
            "dr_aleg_txns": dr_aleg_txns,
        },
    )


def ledger_create(request):
    if request.method == "POST":
        form = LedgerForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Ledger created successfully!")
            return HttpResponse(status=204, headers={"HX-Trigger": "ledgerListChanged"})
    else:
        form = LedgerForm()
    return render(request, "partials/crispy_form.html", {"form": form})


def ledger_update(request, ledger_id):
    ledger = get_object_or_404(Ledger, id=ledger_id)
    if request.method == "POST":
        form = LedgerForm(request.POST, instance=ledger)
        if form.is_valid():
            form.save()
            messages.success(request, "Ledger updated successfully!")
            return HttpResponse(status=204, headers={"HX-Trigger": "ledgerListChanged"})
    else:
        form = LedgerForm(instance=ledger)
    return render(request, "ledger_update.html", {"form": form})


def ledger_save(request, pk=None):
    if pk:
        ledger = get_object_or_404(Ledger, id=pk)
        form = LedgerForm(request.POST or None, instance=ledger)
        verb = "updated"
    else:
        form = LedgerForm(request.POST or None)
        verb = "created"

    if request.method == "POST":
        if form.is_valid():
            form.save()
            messages.success(request, "Ledger {verb} successfully!")
            return HttpResponse(status=204, headers={"HX-Trigger": "listChanged"})

    return render(request, "partials/crispy_form.html", {"form": form})


def ledger_delete(request, ledger_id):
    ledger = get_object_or_404(Ledger, id=ledger_id)
    ledger.delete()
    messages.success(request, "Ledger deleted successfully!")
    return HttpResponse(status=204, headers={"HX-Trigger": "listChanged"})


def ledger_statement_list(request):
    ledger_statements = LedgerStatement.objects.all()
    return render(
        request,
        "dea/ledgerstatement_list.html",
        {"object_list": ledger_statements},
    )


def ledger_statement_detail(request, ledger_statement_id):
    ledger_statement = get_object_or_404(LedgerStatement, id=ledger_statement_id)
    return render(
        request, "ledgerstatement_detail.html", {"ledger_statement": ledger_statement}
    )


def ledger_statement_delete(request, ledger_statement_id):
    ledger_statement = get_object_or_404(LedgerStatement, id=ledger_statement_id)
    ledger_statement.delete()
    return HttpResponse("Ledger statement deleted successfully!")


def ledger_transaction_list(request):
    ledger_transactions = LedgerTransaction.objects.all()
    return render(
        request,
        "ledger_transaction_list.html",
        {"ledger_transactions": ledger_transactions},
    )


def ledger_transaction_detail(request, ledger_transaction_id):
    ledger_transaction = get_object_or_404(LedgerTransaction, id=ledger_transaction_id)
    return render(
        request,
        "ledger_transaction_detail.html",
        {"ledger_transaction": ledger_transaction},
    )


def ledger_transaction_create(request, pk=None):
    journal_entry = get_object_or_404(JournalEntry, pk=pk)
    if request.method == "POST":
        form = LedgerTransactionForm(
            request.POST, initial={"journal_entry": journal_entry}
        )
        if form.is_valid():
            ledger_transaction = form.save(commit=False)
            ledger_transaction.journal_entry = journal_entry
            ledger_transaction.save()
            messages.success(request, "Ledger transaction created successfully!")
            return HttpResponse(status=200, headers={"HX-Trigger": "listChanged"})
    else:
        form = LedgerTransactionForm(
            initial={"journal_entry": journal_entry}, journalentry_id=journal_entry.id
        )
    return render(request, "partials/crispy_form.html", {"form": form})


def ledger_transaction_update(request, pk):
    ledger_transaction = get_object_or_404(LedgerTransaction, id=pk)
    if request.method == "POST":
        form = LedgerTransactionForm(request.POST, instance=ledger_transaction)
        if form.is_valid():
            ledger_transaction = form.save(commit=False)
            ledger_transaction.journal_entry = journal_entry
            ledger_transaction.save()
            messages.success(request, "Ledger transaction updated successfully!")
            return HttpResponse(status=204, headers={"HX-Trigger": "listChanged"})
    else:
        form = LedgerTransactionForm(instance=ledger_transaction)
    return render(request, "partials/crispy_form.html", {"form": form})


def ledger_transaction_delete(request, pk):
    ledger_transaction = get_object_or_404(LedgerTransaction, id=pk)
    ledger_transaction.delete()
    messages.success(request, "Ledger transaction deleted successfully!")
    return HttpResponse(status=204, headers={"HX-Trigger": "listChanged"})


import io
import re

import openpyxl
from django import forms
from django.http import HttpResponse
from django.shortcuts import render
from django.views import View
from faker import Faker


def extract_name(upistring):
    # Check if the input is None or an empty string
    if not upistring:
        return None

    # Define the regex pattern
    pattern = r"UPI-(.*?)-.*?@"

    # Search for the pattern in the input string
    match = re.search(pattern, upistring)

    # If a match is found, return the extracted name
    if match:
        return match.group(1)
    else:
        return None


class UploadFileForm(forms.Form):
    file = forms.FileField()


class FileUploadView(View):
    def get(self, request):
        form = UploadFileForm()
        return render(request, "dea/upload.html", {"form": form})

    #   process gopi sheet
    # def post(self, request):
    #     form = UploadFileForm(request.POST, request.FILES)
    #     if form.is_valid():
    #         file = request.FILES['file']
    #         wb = openpyxl.load_workbook(file)
    #         ws = wb.active

    #         # Initialize Faker
    #         fake = Faker()
    #         # Read each row and generate a list of dictionaries
    #         data = []
    #         for row in ws.iter_rows(min_row=2, values_only=True):  # Skip the header row
    #             voucher_date = row[0]
    #             ledger_name = row[1] if row[1] else fake.name()
    #             voucher_type = row[2]
    #             ledger_amount = row[3]
    #             receipt_ledger = row[4]
    #             ledger_amount_dr_cr = "Cr"  # Assuming Cr for all rows
    #             item_name = row[5]
    #             billed_quantity = row[6]
    #             item_rate = row[7]
    #             # print(voucher_date, ledger_name, voucher_type, ledger_amount, item_name, billed_quantity, item_rate)
    #             if voucher_type == "Sales" or voucher_type == "Purchase" and item_name:
    #                 row_data = [
    #                     voucher_date, voucher_type, ledger_name, ledger_amount, "Dr" if voucher_type =="Sales" else "Cr", "", "", "", "", "", "Item Invoice"
    #                 ]

    #                 second_row_data = [
    #                     "", "", voucher_type, ledger_amount - (ledger_amount * 0.03), "Cr"  if voucher_type =="Sales" else "Dr", item_name, billed_quantity, item_rate, "gms", billed_quantity*item_rate, ""
    #                 ]

    #                 cgst_data = [
    #                     "", "", "Tax - CGST @ 1.5%", ledger_amount * 0.015,"Cr"  if voucher_type =="Sales" else "Dr", "", "", "", "", "", ""
    #                 ]

    #                 sgst_data = [
    #                     "", "",  "Tax - SGST @ 1.5%", ledger_amount * 0.015, "Cr"  if voucher_type =="Sales" else "Dr", "", "", "", "", "", ""
    #                 ]
    #                 data.append(row_data)
    #                 data.append(second_row_data)
    #                 data.append(cgst_data)
    #                 data.append(sgst_data)
    #                 if receipt_ledger and voucher_type == "Sales":
    #                     receipt_ledger_data = [
    #                         voucher_date, "Receipt",ledger_name, ledger_amount, "Cr", "", "", "", "", "", ""
    #                     ]

    #                     receipt_as = [
    #                         "", "", receipt_ledger, ledger_amount, "Dr", "", "", "", "", "", ""
    #                     ]
    #                     data.append(receipt_ledger_data)
    #                     data.append(receipt_as)
    #                 else:
    #                     payment_ledger_data = [
    #                         voucher_date, "Payment",ledger_name, ledger_amount, "Dr", "", "", "", "", "", ""
    #                     ]

    #                     payment_as = [
    #                         "", "", receipt_ledger, ledger_amount, "Cr", "", "", "", "", "", ""
    #                     ]
    #                     data.append(payment_ledger_data)
    #                     data.append(payment_as)
    #         print(data)
    #         # Create a new workbook and worksheet
    #         new_wb = openpyxl.Workbook()
    #         new_ws = new_wb.active
    #         new_ws.title = "Processed Data"

    #         # Define headers
    #         headers = [
    #             "Voucher Date", "Voucher Type Name","Ledger Name", "Ledger Amount", "Ledger Amount Dr/Cr",
    #             "Item Name", "Billed Quantity", "Item Rate", "Item Rate per", "Item Amount", "Change Mode"
    #         ]
    #         new_ws.append(headers)

    #         # Write data to the new worksheet
    #         for row_data in data:
    #             # print(row_data)
    #             new_ws.append(row_data)

    #         # Save the new workbook to a BytesIO object
    #         output = io.BytesIO()
    #         new_wb.save(output)
    #         output.seek(0)

    #         # Return the new Excel file as a response
    #         response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    #         response['Content-Disposition'] = 'attachment; filename="processed_data.xlsx"'
    #         return response
    #     return render(request, 'dea/upload.html', {'form': form})

    def post(self, request):
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES["file"]
            wb = openpyxl.load_workbook(file)
            ws = wb.active

            # Read each row and generate a list of dictionaries
            data = []
            for row in ws.iter_rows(
                min_row=22, values_only=True
            ):  # Skip the header row
                print(row)
                voucher_date = row[0]
                ledger_name = extract_name(row[1])
                refno = row[2]

                withdrawal_amt = row[4]
                deposit_amt = row[5]  # Assuming Cr for all rows
                voucher_type = row[7]
                item_name = row[8]
                item_rate = row[9]
                ledger_amount = (
                    withdrawal_amt
                    if voucher_type in ["Sales", "Payment"]
                    else deposit_amt
                )
                billed_quantity = ledger_amount / item_rate if item_rate else 0
                receipt_ledger = "HDFC_BANK"
                # print(voucher_date, ledger_name, voucher_type, ledger_amount, item_name, billed_quantity, item_rate)
                if voucher_type == "Sales" or voucher_type == "Purchase" and item_name:
                    row_data = [
                        voucher_date,
                        voucher_type,
                        ledger_name,
                        ledger_amount,
                        "Dr" if voucher_type == "Sales" else "Cr",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "Item Invoice",
                    ]

                    second_row_data = [
                        "",
                        "",
                        voucher_type,
                        ledger_amount - (ledger_amount * 0.03),
                        "Cr" if voucher_type == "Sales" else "Dr",
                        item_name,
                        billed_quantity,
                        item_rate,
                        "gms",
                        billed_quantity * item_rate
                        - (billed_quantity * item_rate * 0.3),
                        "",
                    ]

                    cgst_data = [
                        "",
                        "",
                        "Tax - CGST @ 1.5%",
                        ledger_amount * 0.015,
                        "Cr" if voucher_type == "Sales" else "Dr",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                    ]

                    sgst_data = [
                        "",
                        "",
                        "Tax - SGST @ 1.5%",
                        ledger_amount * 0.015,
                        "Cr" if voucher_type == "Sales" else "Dr",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                    ]
                    data.append(row_data)
                    data.append(second_row_data)
                    data.append(cgst_data)
                    data.append(sgst_data)
                    if voucher_type == "Sales":
                        receipt_ledger_data = [
                            voucher_date,
                            "Receipt",
                            ledger_name,
                            ledger_amount,
                            "Cr",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                        ]

                        receipt_as = [
                            "",
                            "",
                            receipt_ledger,
                            ledger_amount,
                            "Dr",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                        ]
                        data.append(receipt_ledger_data)
                        data.append(receipt_as)
                    else:
                        payment_ledger_data = [
                            voucher_date,
                            "Payment",
                            ledger_name,
                            ledger_amount,
                            "Dr",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                        ]

                        payment_as = [
                            "",
                            "",
                            receipt_ledger,
                            ledger_amount,
                            "Cr",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                        ]
                        data.append(payment_ledger_data)
                        data.append(payment_as)
            print(data)
            # Create a new workbook and worksheet
            new_wb = openpyxl.Workbook()
            new_ws = new_wb.active
            new_ws.title = "Processed Data"

            # Define headers
            headers = [
                "Voucher Date",
                "Voucher Type Name",
                "Ledger Name",
                "Ledger Amount",
                "Ledger Amount Dr/Cr",
                "Item Name",
                "Billed Quantity",
                "Item Rate",
                "Item Rate per",
                "Item Amount",
                "Change Mode",
            ]
            new_ws.append(headers)

            # Write data to the new worksheet
            for row_data in data:
                # print(row_data)
                new_ws.append(row_data)

            # Save the new workbook to a BytesIO object
            output = io.BytesIO()
            new_wb.save(output)
            output.seek(0)

            # Return the new Excel file as a response
            response = HttpResponse(
                output,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response[
                "Content-Disposition"
            ] = 'attachment; filename="processed_data.xlsx"'
            return response
        return render(request, "dea/upload.html", {"form": form})
