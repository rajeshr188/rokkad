from datetime import date

from actstream import action
from actstream.models import Action, actor_stream, any_stream, user_stream
from django.contrib.auth.decorators import login_required
from django.db.models import Count, FloatField, Sum
from django.db.models.functions import Cast, Coalesce
from django.http import HttpResponse, HttpResponseForbidden
from django.shortcuts import redirect, render
from django.views.generic import TemplateView

from apps.orgs.decorators import roles_required, workspace_required
from apps.orgs.models import Membership
from apps.tenant_apps.contact.models import Customer
from apps.tenant_apps.contact.services import (active_customers,
                                               get_customers_by_type,
                                               get_customers_by_year)
from apps.tenant_apps.girvi.models import Loan
from apps.tenant_apps.girvi.services import *


class HomePageView(TemplateView):
    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated:

            if request.user.workspace is None or request.user.workspace.schema_name == "public":
                return redirect("dashboard")
            else:
                return redirect("company_dashboard")
        else:
            return super().dispatch(request, *args, **kwargs)

    def get_template_names(self):
        return ["pages/home.html"]


class TenantPageView(TemplateView):
    template_name = "pages/tenant.html"


class AboutPageView(TemplateView):
    template_name = "pages/about.html"


class PrivacyPolicy(TemplateView):
    template_name = "pages/privacy_policy.html"


class CancellationAndRefund(TemplateView):
    template_name = "pages/cancellation_and_refund.html"


class TermsAndConditions(TemplateView):
    template_name = "pages/terms_and_conditions.html"


class ContactPageView(TemplateView):
    template_name = "pages/contact.html"


class HelpPageView(TemplateView):
    template_name = "pages/help.html"


class FaqPageView(TemplateView):
    template_name = "pages/faq.html"


@login_required
# @roles_required(["Owner", "Admin"])
def Dashboard(request):
    context = {}
    return render(request, "pages/dashboard.html", context)


@login_required
@workspace_required
@roles_required(["Owner", "Admin"])
def company_dashboard(request):
    context = {}

    # context['stream'] = user_stream(request.user, with_user_activity=True)
    # context['any_stream'] = any_stream(request.user)
    # context['actor_stream'] = actor_stream(request.user)
    # context['action'] = Action.objects.all()
    company = request.user.workspace
    # try:
    #     membership = Membership.objects.select_related('role').get(user=request.user, company_id=company.id)
    #     if membership.role.name in ["Owner", "Admin"]:
    #         actions = Action.objects.all()[0:10]
    #     else:
    #         actions = user_stream(request.user, with_user_activity=True)[0:10]
    # except Membership.DoesNotExist:
    #     return HttpResponseForbidden()

    # context["actions"] = actions

    # from purchase.models import Invoice as Pinv
    # from sales.models import Invoice as Sinv

    # pinv = Pinv.objects
    # sinv = Sinv.objects
    # total_pbal = pinv.filter(balancetype="Gold").aggregate(
    #     net_wt=Coalesce(Cast(Sum("net_wt"), output_field=FloatField()), 0.0),
    #     gwt=Coalesce(Cast(Sum("gross_wt"), output_field=FloatField()), 0.0),
    #     bal=Coalesce(Cast(Sum("balance"), output_field=FloatField()), 0.0),
    # )
    # total_sbal = sinv.filter(balancetype="Gold").aggregate(
    #     net_wt=Coalesce(Cast(Sum("net_wt"), output_field=FloatField()), 0.0),
    #     gwt=Coalesce(Cast(Sum("gross_wt"), output_field=FloatField()), 0.0),
    #     bal=Coalesce(Cast(Sum("balance"), output_field=FloatField()), 0.0),
    # )
    # total_pbal_ratecut = pinv.filter(balancetype="Cash").aggregate(
    #     net_wt=Coalesce(Cast(Sum("net_wt"), output_field=FloatField()), 0.0),
    #     gwt=Coalesce(Cast(Sum("gross_wt"), output_field=FloatField()), 0.0),
    #     bal=Coalesce(Cast(Sum("balance"), output_field=FloatField()), 0.0),
    # )
    # total_sbal_ratecut = sinv.filter(balancetype="Cash").aggregate(
    #     net_wt=Coalesce(Cast(Sum("net_wt"), output_field=FloatField()), 0.0),
    #     gwt=Coalesce(Cast(Sum("gross_wt"), output_field=FloatField()), 0.0),
    #     bal=Coalesce(Cast(Sum("balance"), output_field=FloatField()), 0.0),
    # )
    # context["total_pbal"] = total_pbal
    # context["total_sbal"] = total_sbal
    # context["pbal"] = total_pbal["bal"] - total_sbal["bal"]
    # context["total_pbal_ratecut"] = total_pbal_ratecut
    # context["total_sbal_ratecut"] = total_sbal_ratecut
    # context["sbal"] = total_pbal_ratecut["bal"] - total_sbal_ratecut["bal"]
    # context["remaining_net_wt"] = (
    #     total_pbal_ratecut["net_wt"] - total_sbal_ratecut["net_wt"]
    # )
    # try:
    #     context["p_map"] = round(
    #         total_pbal_ratecut["bal"] / total_pbal_ratecut["net_wt"], 3
    #     )
    # except ZeroDivisionError:
    #     context["p_map"] = 0.0
    # context['s_map'] = round(total_sbal_ratecut['bal']/total_sbal_ratecut['net_wt'],3)

    context["new_customers"] = Customer.objects.all()[0:5]
    context["customer_count"] = Customer.objects.values("customer_type").annotate(
        count=Count("id")
    )
    context["grouped_loan_counts"] = get_loan_counts_grouped()
    loan = Loan.objects.with_details(grate=request.grate, srate=request.srate)
    released = loan.released()
    unreleased = loan.unreleased()
    sunken = unreleased.filter(is_overdue="True")
    today = date.today()
    today_loan = LoanItem.objects.filter(loan__loan_date__gte=today).aggregate(
        amount=Sum("loanamount"), interest=Sum("interest")
    )
    today_release = Release.objects.filter(release_date__gte=today).aggregate(
        amount=Sum("loan__loan_amount"), interest=Sum("loan__interest")
    )
    context["today_loan"] = today_loan
    context["loan_count"] = unreleased.count()

    context["due_amount"] = unreleased.aggregate(
        Sum("loan_amount"), Sum("total_interest"), Sum("total_due")
    )
    context["total_loan_amount"] = context["due_amount"]["loan_amount__sum"]
    context["total_interest"] = context["due_amount"]["total_interest__sum"]

    context[
        "assets"
    ] = unreleased.with_itemwise_loanamount().total_itemwise_loanamount()
    context["loanbyitemtype"] = get_loanamount_by_itemtype()
    context["weight"] = unreleased.total_weight()
    context["pure_weight"] = unreleased.total_pure_weight()

    context["current_value"] = unreleased.total_current_value()
    context["itemwise_value"] = unreleased.itemwise_value()
    context["total_current_value"] = unreleased.total_current_value()["total"]

    context["sunken"] = {}
    context["sunken"]["loan_count"] = sunken.count()
    context["sunken"]["total_loan_amount"] = sunken.total_loanamount()
    context["sunken"][
        "assets"
    ] = sunken.with_itemwise_loanamount().total_itemwise_loanamount()
    context["sunken"]["weight"] = sunken.total_weight()
    context["sunken"]["due_amount"] = sunken.aggregate(
        Sum("loan_amount"), Sum("total_interest"), Sum("total_due")
    )
    context["sunken"]["current_value"] = sunken.total_current_value()
    context["sunken"]["itemwise_value"] = sunken.itemwise_value()
    context["sunken"]["total_current_value"] = sunken.total_current_value()["total"]
    context["sunken"]["total_interest"] = sunken.aggregate(total=Sum("total_interest"))
    context["sunken"]["pure_weight"] = sunken.total_pure_weight()

    try:
        context["loan_progress"] = round(
            loan.released().count() / loan.count() * 100, 2
        )
    except ZeroDivisionError:
        context["loan_progress"] = 0.0
    context["loan_data_by_year"] = get_loans_by_year()
    context["customer_data_by_year"] = get_customers_by_year()
    context["customer_data_by_type"] = get_customers_by_type()
    context["active_customers"] = active_customers()
    context["avg_loan_per_day"] = get_average_loan_instance_per_day()
    context["maxloans"] = (
        Customer.objects.filter(loan__release__isnull=True)
        .annotate(
            num_loans=Count("loan"),
            sum_loans=Sum("loan__loan_amount"),
            tint=Sum("loan__interest"),
        )
        .values("name", "num_loans", "sum_loans", "tint")
        .order_by("-num_loans", "sum_loans", "tint")
    )
    context["interest_received"] = get_interest_paid()
    context["loan_cumsum"] = list(get_loan_cumulative_amount())
    return render(request, "pages/company_dashboard.html", context)


from django.apps import apps
from django.db import transaction
from moneyed import Money
from openpyxl import load_workbook

from apps.tenant_apps.contact.models import Customer
from apps.tenant_apps.dea.models import (AccountStatement, AccountTransaction,
                                         JournalEntry, Ledger,
                                         LedgerTransaction, TransactionType_DE,
                                         TransactionType_Ext)
from apps.tenant_apps.dea.utils.currency import Balance
from apps.tenant_apps.purchase.models import Payment, Purchase
from apps.tenant_apps.sales.models import Invoice, Receipt

from .forms import MaxxFileUploadForm


def create_purchase_data(file):
    wb = load_workbook(file)
    sheet = wb.active
    # List of headers you want
    headers = [
        "Invoice No",
        "Vou Date",
        "Account Name",
        "Amount",
        "Chrgd.Wgt",
        "Pure Balance",
    ]  # replace with your headers
    # Get the second row values (headers in the Excel file)
    excel_headers = [cell.value for cell in sheet[2]]
    # Get the index of each header in your list
    header_indices = {header: excel_headers.index(header) for header in headers}

    purchases_data = []
    for row in sheet.iter_rows(min_row=5, values_only=True):
        if all(
            cell is None for cell in row
        ):  # Stop reading when an empty row is encountered
            break
        values = {header: row[index] for header, index in header_indices.items()}

        supplier, created = Customer.objects.get_or_create(name=values["Account Name"])
        purchases_data.append(
            Purchase(
                voucher_date=values["Vou Date"],
                voucher_no=values["Invoice No"],
                supplier=supplier,
                balance_gold=Money(values["Pure Balance"], "USD"),
                balance_cash=Money(values["Amount"], "INR"),
            )
        )  # replace field1, field2, field3 with your actual field names
    instances = Purchase.objects.bulk_create(purchases_data)
    return instances


def create_sales_data(file):
    wb = load_workbook(file)
    sheet = wb.active
    # List of headers you want
    headers = [
        "Invoice No",
        "Vou Date",
        "Account Name",
        "Amount",
        "Chrgd.Wgt",
        "Pure Balance",
    ]  # replace with your headers
    # Get the second row values (headers in the Excel file)
    excel_headers = [cell.value for cell in sheet[2]]
    # Get the index of each header in your list
    header_indices = {header: excel_headers.index(header) for header in headers}

    sales_data = []
    for row in sheet.iter_rows(min_row=5, values_only=True):
        if all(
            cell is None for cell in row
        ):  # Stop reading when an empty row is encountered
            break
        values = {header: row[index] for header, index in header_indices.items()}
        # {'Vou Date': datetime.datetime(2024, 4, 14, 0, 0), 'Account Name': 'Karuna Padaved', 'Amount': 0, 'Chrgd.Wgt': 6.674}

        customer, created = Customer.objects.get_or_create(name=values["Account Name"])
        balance_cash = Money(values["Amount"] or 0, "INR")
        balance_gold = Money(values["Pure Balance"] or 0, "USD")
        sales_data.append(
            Invoice(
                voucher_date=values["Vou Date"],
                voucher_no=values["Invoice No"],
                customer=customer,
                balance_cash=balance_cash,
                balance_gold=balance_gold,
            )
        )  # replace field1, field2, field3 with your actual field names
    instances = Invoice.objects.bulk_create(sales_data)
    return instances


def create_payment_data(file):
    wb = load_workbook(file)
    sheet = wb.active
    # List of headers you want
    headers = [
        "Ref No",
        "Vou Date",
        "Account Name",
        "Gold",
        "Silver",
        "Rate",
        "Conversion Value",
        "Amount",
    ]  # replace with your headers
    # Get the second row values (headers in the Excel file)
    excel_headers = [cell.value for cell in sheet[2]]
    # Get the index of each header in your list
    header_indices = {header: excel_headers.index(header) for header in headers}
    payment_data = []
    for row in sheet.iter_rows(min_row=4, values_only=True):
        if all(
            cell is None for cell in row
        ):  # Stop reading when an empty row is encountered
            break
        values = {header: row[index] for header, index in header_indices.items()}
        supplier, created = Customer.objects.get_or_create(name=values["Account Name"])

        if values["Gold"] is not None:
            total = Money(values["Gold"], "USD")
        elif values["Amount"] is not None:
            total = Money(values["Amount"], "INR")
        else:
            total = Money(0, "USD")

        payment_data.append(
            Payment(
                voucher_date=values["Vou Date"],
                voucher_no=values["Ref No"],
                supplier=supplier,
                total=total,
            )
        )  # replace field1, field2, field3 with your actual field names
    instances = Payment.objects.bulk_create(payment_data)
    return instances


def create_receipt_data(file):
    wb = load_workbook(file)
    sheet = wb.active
    # List of headers you want
    headers = [
        "Ref No",
        "Vou Date",
        "Account Name",
        "Gold",
        "Silver",
        "Rate",
        "Conversion Value",
        "Amount",
    ]  # replace with your headers
    # Get the second row values (headers in the Excel file)
    excel_headers = [cell.value for cell in sheet[2]]
    # Get the index of each header in your list
    header_indices = {header: excel_headers.index(header) for header in headers}
    receipt_data = []
    for row in sheet.iter_rows(min_row=4, values_only=True):
        if all(
            cell is None for cell in row
        ):  # Stop reading when an empty row is encountered
            break
        values = {header: row[index] for header, index in header_indices.items()}

        customer, created = Customer.objects.get_or_create(name=values["Account Name"])
        if values["Conversion Value"] is not None:
            weight = values["Gold"]
            rate = values["Rate"]
            total = Money(values["Gold"], "USD")
            receipt = Receipt(
                voucher_date=values["Vou Date"],
                voucher_no=values["Ref No"],
                customer=customer,
                weight=weight,
                touch=100,
                rate=rate,
                convert=True,
                amount=Money(values["Amount"] or 0, "INR"),
                total=total,
            )
        else:
            if values["Amount"] is not None:
                total = Money(values["Amount"], "INR")
            elif values["Gold"] is not None:
                total = Money(values["Gold"], "USD")
            receipt = Receipt(
                voucher_date=values["Vou Date"],
                voucher_no=values["Ref No"],
                customer=customer,
                total=total,
            )
        receipt_data.append(
            receipt
        )  # replace field1, field2, field3 with your actual field names
    instances = Receipt.objects.bulk_create(receipt_data)
    return instances


def create_debtors_data(file):
    wb = load_workbook(file)
    sheet = wb.active

    account_data = []
    # Fetch all accounts and store them in a dictionary
    # accounts = {account.name: account for account in Account.objects.all()}
    for row in sheet.iter_rows(min_row=9, values_only=True):
        if all(
            cell is None for cell in row
        ):  # Stop reading when an empty row is encountered
            break

        name = row[1]  # Column B is the 2nd column, but Python uses 0-based indexing
        customer, created = Customer.objects.get_or_create(name=name, customer_type="W")

        if not hasattr(customer, "account"):
            customer.save()

        # cash_debit = row[5]  # Column C
        # cash_credit = row[6]  # Column D
        # gold_debit = row[7]  # Column E
        # gold_credit = row[8]  # Column F
        cash_debit = row[7]  # Column C
        cash_credit = row[15]  # Column D
        gold_debit = row[21]  # Column E
        gold_credit = row[23]  # Column F
        print(customer, cash_debit, cash_credit, gold_debit, gold_credit)
        if cash_credit is not None:
            cash = Money(-cash_credit, "INR")
        elif cash_debit is not None:
            cash = Money(cash_debit, "INR")
        else:
            cash = Money(0, "INR")
        if gold_credit is not None:
            gold = Money(-gold_credit, "USD")
        elif gold_debit is not None:
            gold = Money(gold_debit, "USD")
        else:
            gold = Money(0, "USD")
        balance = Balance([cash, gold])
        tc = Balance(0, "INR")
        td = Balance(0, "INR")
        account_data.append(
            AccountStatement(
                AccountNo=customer.account,
                ClosingBalance=balance.monies(),
                TotalCredit=tc.monies(),
                TotalDebit=td.monies(),
            )
        )  # replace field1, field2, field3 with your actual field names
    AccountStatement.objects.bulk_create(account_data)


def create_creditors_data(file):
    wb = load_workbook(file)
    sheet = wb.active

    account_data = []
    # Fetch all accounts and store them in a dictionary
    # accounts = {account.name: account for account in Account.objects.all()}
    for row in sheet.iter_rows(min_row=4, values_only=True):
        if all(
            cell is None for cell in row
        ):  # Stop reading when an empty row is encountered
            break

        name = row[1]  # Column B is the 2nd column, but Python uses 0-based indexing
        customer, created = Customer.objects.get_or_create(name=name, customer_type="S")
        if not hasattr(customer, "account"):
            customer.save()

        cash_debit = row[5]  # Column C
        cash_credit = row[6]  # Column D
        gold_debit = row[7]  # Column E
        gold_credit = row[8]  # Column F
        # print(cash_debit,cash_credit,gold_debit,gold_credit)
        if cash_credit is not None:
            cash = Money(cash_credit, "INR")
        elif cash_debit is not None:
            cash = Money(-cash_debit, "INR")
        else:
            cash = Money(0, "INR")
        if gold_credit is not None:
            gold = Money(gold_credit, "USD")
        elif gold_debit is not None:
            gold = Money(-gold_debit, "USD")
        else:
            gold = Money(0, "USD")
        balance = Balance([cash, gold])
        tc = Balance(0, "INR")
        td = Balance(0, "INR")
        account_data.append(
            AccountStatement(
                AccountNo=customer.account,
                ClosingBalance=balance.monies(),
                TotalCredit=tc.monies(),
                TotalDebit=td.monies(),
            )
        )  # replace field1, field2, field3 with your actual field names
    AccountStatement.objects.bulk_create(account_data)


# def handle_import(data,model):
#     wb = load_workbook(file)
#     sheet = wb.active

#     purchases_data = []
#     for row in sheet.iter_rows(min_row=5, values_only=True):
#         if all(cell is None for cell in row):  # Stop reading when an empty row is encountered
#             break

#         date = row[6]  # Column G is the 7th column, but Python uses 0-based indexing
#         supplier = row[7]  # Column H
#         balance_gold = row[16]  # Column Q

#         supplier,created = Customer.objects.get_or_create(name=supplier, customer_type="S")
#         balance_gold = Money(balance_gold, 'USD')
#         purchases_data.append(Purchase(supplier=supplier, balance_gold = balance_gold))  # replace field1, field2, field3 with your actual field names


#     with transaction.atomic():
#         # Step 1: Create a list of JournalEntry instances
#         purchases = Purchase.objects.bulk_create(purchases_data)
#         journal_entries = [JournalEntry(content_object=purchase) for purchase in purchases]

#         # Step 2: Bulk create the JournalEntry instances
#         created_journal_entries = JournalEntry.objects.bulk_create(journal_entries)

#         ledger_transactions = []
#         account_transactions = []

#         # Create a dictionary to cache Ledger instances
#         ledgers = {}

#         # Step 3: For each JournalEntry instance, get the LedgerTransaction and AccountTransaction data
#         for journal_entry, purchase in zip(created_journal_entries, purchases):
#             lt_data, at_data = purchase.get_transactions()

#             for lt in lt_data:
#                 # Fetch the Ledger instance from the database if it's not already in the cache
#                 for key in ['ledgerno', 'ledgerno_dr']:
#                     if lt[key] not in ledgers:
#                         ledgers[lt[key]] = Ledger.objects.get(name=lt[key])
#                     lt[key] = ledgers[lt[key]]

#                 ledger_transaction = LedgerTransaction(journal_entry=journal_entry, **lt)
#                 ledger_transactions.append(ledger_transaction)

#             for at in at_data:
#                  # Fetch the Ledger instance from the database if it's not already in the cache
#                 if at['ledgerno'] not in ledgers:
#                     ledgers[at['ledgerno']] = Ledger.objects.get(name=at['ledgerno'])
#                 at['ledgerno'] = ledgers[at['ledgerno']]

#                 at["XactTypeCode"] = TransactionType_DE.objects.get(XactTypeCode=at["XactTypeCode"])
#                 at["XactTypeCode_ext"] = TransactionType_Ext.objects.get(XactTypeCode_ext=at["XactTypeCode_ext"])
#                 account_transaction = AccountTransaction(journal_entry=journal_entry, **at)
#                 account_transactions.append(account_transaction)

#         # Step 4: Bulk create the LedgerTransaction and AccountTransaction instances
#         lts = LedgerTransaction.objects.bulk_create(ledger_transactions)
#         ats = AccountTransaction.objects.bulk_create(account_transactions)
#         print(f"lts:{len(lts)} , ats:{len(ats)}")


def handle_import(instances):
    with transaction.atomic():
        journal_entries = [
            JournalEntry(content_object=instance) for instance in instances
        ]
        created_journal_entries = JournalEntry.objects.bulk_create(journal_entries)
        ledger_transactions = []
        account_transactions = []
        ledgers = {}
        for journal_entry, instance in zip(created_journal_entries, instances):
            lt_data, at_data = instance.get_transactions()
            # print(lt_data,at_data)
            for lt in lt_data:
                for key in ["ledgerno", "ledgerno_dr"]:
                    if lt[key] not in ledgers:
                        ledgers[lt[key]] = Ledger.objects.get(name=lt[key])
                    lt[key] = ledgers[lt[key]]
                ledger_transaction = LedgerTransaction(
                    journal_entry=journal_entry, **lt
                )
                ledger_transactions.append(ledger_transaction)
            for at in at_data:
                if at["ledgerno"] not in ledgers:
                    ledgers[at["ledgerno"]] = Ledger.objects.get(name=at["ledgerno"])
                at["ledgerno"] = ledgers[at["ledgerno"]]
                at["XactTypeCode"] = TransactionType_DE.objects.get(
                    XactTypeCode=at["XactTypeCode"]
                )
                at["XactTypeCode_ext"] = TransactionType_Ext.objects.get(
                    XactTypeCode_ext=at["XactTypeCode_ext"]
                )
                account_transaction = AccountTransaction(
                    journal_entry=journal_entry, **at
                )
                account_transactions.append(account_transaction)
        LedgerTransaction.objects.bulk_create(ledger_transactions)
        AccountTransaction.objects.bulk_create(account_transactions)


def maxx_files_upload(request):
    if request.method == "POST":
        form = MaxxFileUploadForm(request.POST, request.FILES)
        if form.is_valid():
            model = form.cleaned_data["model"]
            file = form.cleaned_data["file"]
            if model == "Purchase":
                data = create_purchase_data(file)
                handle_import(data)
            elif model == "Sales":
                data = create_sales_data(file)
                handle_import(data)
            elif model == "Payment":
                data = create_payment_data(file)
                handle_import(data)
            elif model == "Receipt":
                data = create_receipt_data(file)
                handle_import(data)
            elif model == "Debtors":
                create_debtors_data(file)
            elif model == "Creditors":
                create_creditors_data(file)

            return HttpResponse("File uploaded & imported successfully")
    else:
        form = MaxxFileUploadForm()
    return render(request, "pages/maxx_files_upload.html", {"form": form})


# def handle_import(file, model, field_mapping):
#     wb = load_workbook(file)
#     sheet = wb.active
#     data = []
#     for row in sheet.iter_rows(min_row=5, values_only=True):
#         if all(cell is None for cell in row):  # Stop reading when an empty row is encountered
#             break
#         row_data = {}
#         for field, column in field_mapping.items():
#             row_data[field] = row[column]
#         data.append(model(**row_data))

#     with transaction.atomic():
#         instances = model.objects.bulk_create(data)
#         journal_entries = [JournalEntry(content_object=instance) for instance in instances]
#         created_journal_entries = JournalEntry.objects.bulk_create(journal_entries)
#         ledger_transactions = []
#         account_transactions = []
#         ledgers = {}
#         for journal_entry, instance in zip(created_journal_entries, instances):
#             lt_data, at_data = instance.get_transactions()
#             for lt in lt_data:
#                 for key in ['ledgerno', 'ledgerno_dr']:
#                     if lt[key] not in ledgers:
#                         ledgers[lt[key]] = Ledger.objects.get(name=lt[key])
#                     lt[key] = ledgers[lt[key]]
#                 ledger_transaction = LedgerTransaction(journal_entry=journal_entry, **lt)
#                 ledger_transactions.append(ledger_transaction)
#             for at in at_data:
#                 if at['ledgerno'] not in ledgers:
#                     ledgers[at['ledgerno']] = Ledger.objects.get(name=at['ledgerno'])
#                 at['ledgerno'] = ledgers[at['ledgerno']]
#                 at["XactTypeCode"] = TransactionType_DE.objects.get(XactTypeCode=at["XactTypeCode"])
#                 at["XactTypeCode_ext"] = TransactionType_Ext.objects.get(XactTypeCode_ext=at["XactTypeCode_ext"])
#                 account_transaction = AccountTransaction(journal_entry=journal_entry, **at)
#                 account_transactions.append(account_transaction)
#         lts = LedgerTransaction.objects.bulk_create(ledger_transactions)
#         ats = AccountTransaction.objects.bulk_create(account_transactions)
#         print(f"lts:{len(lts)} , ats:{len(ats)}")
